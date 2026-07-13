"""Excel workbook extraction utilities."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from .file_utils import make_report, now_iso, safe_sheet_filename, write_csv, write_json, write_text
from .paths import relative_path


def cell_value_for_csv(value: Any) -> Any:
    """Return a CSV-safe cell value."""
    if value is None:
        return ""
    return value


def extract_excel_to_parsed(
    source_path: Path,
    *,
    parsed_dir: Path,
    tables_dir: Path,
    document_path: Path,
    metadata_path: Path,
    report_path: Path,
    metadata_base: dict[str, Any],
) -> dict[str, Any]:
    """Extract an Excel workbook into CSV tables, metadata, and a Markdown summary."""
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - depends on environment
        message = f"openpyxl is not installed or failed to import: {exc}"
        return write_excel_error(metadata_base, document_path, metadata_path, report_path, message)

    try:
        workbook = load_workbook(source_path, data_only=False, read_only=False)
    except Exception as exc:
        return write_excel_error(metadata_base, document_path, metadata_path, report_path, f"Failed to open workbook: {exc}")

    sheets: list[dict[str, Any]] = []
    markdown_lines = [
        f"# {metadata_base['title']}",
        "",
        "## Workbook Summary",
        "",
        f"- original_filename: {metadata_base['original_filename']}",
        f"- file_type: {metadata_base['file_type']}",
        f"- sheet_count: {len(workbook.worksheets)}",
        "",
        "## Sheets",
        "",
    ]

    for sheet in workbook.worksheets:
        csv_name = safe_sheet_filename(sheet.title)
        csv_path = tables_dir / csv_name
        rows = []
        non_empty_cells = 0
        formula_cells = 0
        for row in sheet.iter_rows():
            csv_row = []
            for cell in row:
                value = cell.value
                if value not in (None, ""):
                    non_empty_cells += 1
                if isinstance(value, str) and value.startswith("="):
                    formula_cells += 1
                csv_row.append(cell_value_for_csv(value))
            rows.append(csv_row)
        write_csv(csv_path, rows)
        merged_ranges = [str(item) for item in sheet.merged_cells.ranges]
        sheet_info = {
            "name": sheet.title,
            "csv_path": relative_path(csv_path),
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "non_empty_cells": non_empty_cells,
            "formula_cells": formula_cells,
            "merged_ranges": merged_ranges,
        }
        sheets.append(sheet_info)
        markdown_lines.extend(
            [
                f"### {sheet.title}",
                "",
                f"- csv_path: `{sheet_info['csv_path']}`",
                f"- rows: {sheet.max_row}",
                f"- columns: {sheet.max_column}",
                f"- non_empty_cells: {non_empty_cells}",
                f"- formula_cells: {formula_cells}",
                f"- merged_ranges: {len(merged_ranges)}",
                "",
            ]
        )

    metadata = {
        **metadata_base,
        "status": "ok",
        "converter": "openpyxl",
        "processed_at": now_iso(),
        "sheets": sheets,
    }
    write_text(document_path, "\n".join(markdown_lines))
    write_json(metadata_path, metadata)
    write_text(report_path, make_report(metadata_base["title"], "ok", ["- converter: openpyxl", f"- parsed_dir: {relative_path(parsed_dir)}"]))
    return {
        "status": "ok",
        "converter": "openpyxl",
        "document_path": relative_path(document_path),
        "metadata_path": relative_path(metadata_path),
        "report_path": relative_path(report_path),
        "sheets": sheets,
    }


def write_excel_error(
    metadata_base: dict[str, Any],
    document_path: Path,
    metadata_path: Path,
    report_path: Path,
    message: str,
) -> dict[str, Any]:
    """Write error metadata/report for an Excel extraction failure."""
    metadata = {
        **metadata_base,
        "status": "error",
        "converter": "openpyxl",
        "processed_at": now_iso(),
        "error": message,
        "sheets": [],
    }
    write_text(document_path, f"# {metadata_base['title']}\n\nExtraction failed.\n")
    write_json(metadata_path, metadata)
    write_text(report_path, make_report(metadata_base["title"], "error", [f"- error: {message}"]))
    return {
        "status": "error",
        "converter": "openpyxl",
        "error": message,
    }

